import datetime
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

from django.shortcuts import render
from django.views import View
from openpyxl import load_workbook

from ..models import PlanningBatch, PlanningEntry


def _normalize_header(name: str) -> str:
    decomposed = unicodedata.normalize("NFKD", name)
    ascii_only = decomposed.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-z0-9]+", "_", ascii_only.strip().lower())
    return normalized.strip("_")


def _index_map(headers) -> Dict[str, int]:
    return {_normalize_header(str(h)): idx for idx, h in enumerate(headers) if h}


def _pick(indexes: Dict[str, int], *candidates: str) -> Optional[int]:
    for key in candidates:
        if key in indexes:
            return indexes[key]
    return None


def _parse_decimal(value: Any, precision: int = 2) -> Optional[Decimal]:
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
        if precision is not None:
            quant = Decimal("1").scaleb(-precision)
            number = number.quantize(quant)
        return number
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(str(value).split(".")[0])
    except (ValueError, TypeError):
        return None


def _parse_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in {"si", "sí", "true", "1", "x", "yes"}


def _parse_date(value: Any) -> Optional[datetime.date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _value(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _preview_rows(rows: List[List[Any]], limit: int = 10) -> List[List[Any]]:
    return [list(r) for r in rows[:limit]]


class PlanningUploadView(View):
    template_name = "planning_upload.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        errors: List[str] = []
        plan_date_raw = request.POST.get("plan_date")
        selected_sheet = request.POST.get("sheet_name") or ""
        step = request.POST.get("step") or "preview"
        upload = request.FILES.get("file")

        plan_date = None
        if plan_date_raw:
            try:
                plan_date = datetime.date.fromisoformat(plan_date_raw)
            except ValueError:
                errors.append("Fecha inválida. Usa formato AAAA-MM-DD.")
        else:
            errors.append("Debes seleccionar una fecha de planificación.")

        if not upload:
            errors.append("Debes subir un archivo Excel (.xlsx).")

        sheet_names: List[str] = []
        preview_headers: List[str] = []
        preview_rows: List[List[Any]] = []
        summary: Dict[str, Any] = {}

        if errors:
            return render(
                request,
                self.template_name,
                {"errors": errors, "plan_date": plan_date_raw},
            )

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception as exc:
            errors.append(f"No se pudo leer el Excel: {exc}")
            return render(request, self.template_name, {"errors": errors, "plan_date": plan_date_raw})

        sheet_names = workbook.sheetnames
        if not sheet_names:
            errors.append("El archivo no tiene hojas (libros).")
            return render(request, self.template_name, {"errors": errors, "plan_date": plan_date_raw})

        if selected_sheet not in sheet_names:
            selected_sheet = sheet_names[0]

        sheet = workbook[selected_sheet]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            errors.append("La hoja seleccionada no tiene filas.")
            return render(request, self.template_name, {"errors": errors, "plan_date": plan_date_raw})

        headers = rows[0]
        preview_headers = [str(h) if h is not None else "" for h in headers]
        indexes = _index_map(headers)
        body_rows = rows[1:]
        preview_rows = _preview_rows(body_rows)
        if step != "import":
            summary = {
                "sheet_name": selected_sheet,
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "filename": upload.name,
                "pending_import": True,
            }
            return render(
                request,
                self.template_name,
                {
                    "plan_date": plan_date_raw,
                    "sheet_names": sheet_names,
                    "sheet_count": len(sheet_names),
                    "selected_sheet": selected_sheet,
                    "preview_headers": preview_headers,
                    "preview_rows": preview_rows,
                    "summary": summary,
                },
            )

        batch = PlanningBatch.objects.create(
            plan_date=plan_date,
            sheet_name=selected_sheet,
            source_filename=upload.name,
        )

        created = 0
        skipped = 0

        for row in body_rows:
            if all(cell in (None, "") for cell in row):
                skipped += 1
                continue

            entry = PlanningEntry(
                batch=batch,
                external_id=str(_value(row, _pick(indexes, "id")) or ""),
                tipo_carga=str(_value(row, _pick(indexes, "tipo_carga", "tipocarga")) or ""),
                ranking_tienda=_parse_int(_value(row, _pick(indexes, "ranking_tienda", "ranking", "rankingt"))),
                sucursal=str(_value(row, _pick(indexes, "sucursal", "tienda")) or ""),
                item_code=str(_value(row, _pick(indexes, "item_code", "itemcode", "codigo")) or ""),
                item_name=str(_value(row, _pick(indexes, "item_name", "itemname", "descripcion")) or ""),
                u_categoria=str(_value(row, _pick(indexes, "u_categoria", "u_categoria")) or ""),
                categoria=str(_value(row, _pick(indexes, "categoria")) or ""),
                a_despachar_total=_parse_decimal(_value(row, _pick(indexes, "a_despachar_total", "a_despachar"))),
                motivo_decision=str(_value(row, _pick(indexes, "motivo_decision", "motivo")) or ""),
                en_transito=_parse_decimal(_value(row, _pick(indexes, "en_transito"))),
                ult_entrada_almacen=_parse_date(_value(row, _pick(indexes, "ult_entrada_almacen", "ult_entrada"))),
                ult_venta_tienda=_parse_date(_value(row, _pick(indexes, "ult_venta_tienda", "ult_venta"))),
                dias_permanencia=_parse_int(_value(row, _pick(indexes, "dias_permanencia"))),
                venta_diaria=_parse_decimal(_value(row, _pick(indexes, "venta_diaria"))),
                stock_tienda=_parse_decimal(_value(row, _pick(indexes, "stock_tienda"))),
                stock_cedis=_parse_decimal(_value(row, _pick(indexes, "stock_cedis", "stock_cedis"))),
                necesidad_urgente=_parse_bool(_value(row, _pick(indexes, "necesidad_urgente", "urgente"))),
                origen_picking=str(_value(row, _pick(indexes, "origen_picking", "origen")) or ""),
                no_planificar=_parse_bool(_value(row, _pick(indexes, "no_planificar", "no_planificar"))),
                ng=_parse_decimal(_value(row, _pick(indexes, "ng"))),
                ccct=_parse_decimal(_value(row, _pick(indexes, "ccct"))),
                sm=_parse_decimal(_value(row, _pick(indexes, "sm"))),
                cubicaje_unidad=_parse_decimal(_value(row, _pick(indexes, "cubicaje_unidad")), precision=4),
                cubicaje_total=_parse_decimal(_value(row, _pick(indexes, "cubicaje_total")), precision=4),
            )
            entry.save()
            created += 1

        summary = {
            "created_entries": created,
            "skipped_rows": skipped,
            "sheet_name": selected_sheet,
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "filename": upload.name,
            "batch_id": batch.id,
        }

        return render(
            request,
            self.template_name,
            {
                "plan_date": plan_date_raw,
                "sheet_names": sheet_names,
                "sheet_count": len(sheet_names),
                "selected_sheet": selected_sheet,
                "preview_headers": preview_headers,
                "preview_rows": preview_rows,
                "summary": summary,
            },
        )
