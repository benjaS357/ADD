import datetime
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional

from django.shortcuts import render
from django.views import View
from openpyxl import load_workbook

from ..models import Salida


def _normalize_header(name: str) -> str:
    decomposed = unicodedata.normalize("NFKD", name)
    ascii_only = decomposed.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-z0-9]+", "_", ascii_only.strip().lower())
    return normalized.strip("_")


def _index_map(headers) -> Dict[str, int]:
    return {_normalize_header(str(h)): idx for idx, h in enumerate(headers) if h}


def _pick(indexes: Dict[str, int], *candidates: str) -> Optional[int]:
    for key in candidates:
        if key in indexes:
            return indexes[key]
        # allow camelCase headers collapsed without separators
        collapsed = key.replace("_", "")
        if collapsed in indexes:
            return indexes[collapsed]
    return None


def _value(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_decimal(value: Any) -> Optional[Decimal]:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _parse_date(value: Any) -> Optional[datetime.date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _preview_rows(rows: List[List[Any]], limit: int = 10) -> List[List[Any]]:
    return [list(r) for r in rows[:limit]]


class SalidaUploadView(View):
    template_name = "salida_upload.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        errors: List[str] = []
        step = request.POST.get("step") or "preview"
        upload = request.FILES.get("file")

        if not upload:
            errors.append("Debes subir un archivo Excel (.xlsx).")
            return render(request, self.template_name, {"errors": errors})

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception as exc:
            errors.append(f"No se pudo leer el Excel: {exc}")
            return render(request, self.template_name, {"errors": errors})

        sheet_names = workbook.sheetnames
        if not sheet_names:
            errors.append("El archivo no tiene hojas.")
            return render(request, self.template_name, {"errors": errors})

        sheet = workbook[sheet_names[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            errors.append("La hoja no tiene filas.")
            return render(request, self.template_name, {"errors": errors})

        headers = rows[0]
        preview_headers = [str(h) if h is not None else "" for h in headers]
        indexes = _index_map(headers)
        body_rows = rows[1:]
        preview_rows = _preview_rows(body_rows)

        if step != "import":
            summary = {
                "sheet_name": sheet.title,
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "filename": upload.name,
                "pending_import": True,
            }
            return render(
                request,
                self.template_name,
                {
                    "sheet_names": sheet_names,
                    "selected_sheet": sheet.title,
                    "preview_headers": preview_headers,
                    "preview_rows": preview_rows,
                    "summary": summary,
                },
            )

        # Detect target month/year from first valid fecha_salida (fallback: fecha_entrada)
        target_date = None
        for row in body_rows:
            fecha_salida = _parse_date(_value(row, _pick(indexes, "fecha_salida", "fecha_salida")))
            fecha_entrada = _parse_date(_value(row, _pick(indexes, "fecha_entrada", "fecha_entrad")))
            target_date = fecha_salida or fecha_entrada
            if target_date:
                break

        if not target_date:
            errors.append("No se encontró una fecha en la hoja (columna Fecha Salida o Fecha_Entrada).")
            return render(
                request,
                self.template_name,
                {
                    "errors": errors,
                    "sheet_names": sheet_names,
                    "selected_sheet": sheet.title,
                    "preview_headers": preview_headers,
                    "preview_rows": preview_rows,
                },
            )

        target_year, target_month = target_date.year, target_date.month

        created = 0
        skipped = 0
        updated = 0
        skipped_wrong_month = 0
        skipped_no_date = 0

        for row in body_rows:
            if all(cell in (None, "") for cell in row):
                skipped += 1
                continue

            fecha_salida = _parse_date(_value(row, _pick(indexes, "fecha_salida", "fecha_salida")))
            fecha_entrada = _parse_date(_value(row, _pick(indexes, "fecha_entrada", "fecha_entrad")))
            row_date = fecha_salida or fecha_entrada
            if not row_date:
                skipped_no_date += 1
                continue
            if row_date.year != target_year or row_date.month != target_month:
                skipped_wrong_month += 1
                continue

            sku = str(_value(row, _pick(indexes, "sku")) or "")
            salida = str(_value(row, _pick(indexes, "salida")) or "")

            defaults = {
                "nombre_sucursal_origen": str(_value(row, _pick(indexes, "nombre_sucursal_origen", "sucursal_origen")) or ""),
                "nombre_almacen_origen": str(_value(row, _pick(indexes, "nombre_almacen_origen", "almacen_origen", "nombrealmacenorigen")) or ""),
                "descripcion": str(_value(row, _pick(indexes, "descripcion")) or ""),
                "cantidad": _parse_decimal(_value(row, _pick(indexes, "cantidad"))),
                "sucursal_destino_propuesto": str(_value(row, _pick(indexes, "sucursal_destino_propuesto", "sucursal_destino")) or ""),
                "entrada": str(_value(row, _pick(indexes, "entrada")) or ""),
                "fecha_entrada": fecha_entrada,
                "nombre_sucursal_destino": str(_value(row, _pick(indexes, "nombre_sucursal_destino", "sucursal_destino", "nombresucursaldestino")) or ""),
                "nombre_almacen_destino": str(_value(row, _pick(indexes, "nombre_almacen_destino", "almacen_destino", "nombrealmacendestino")) or ""),
                "comments": str(_value(row, _pick(indexes, "comments", "comentarios", "comentario")) or ""),
            }

            obj, created_flag = Salida.objects.update_or_create(
                sku=sku,
                fecha_salida=row_date,
                salida=salida,
                defaults=defaults,
            )
            if created_flag:
                created += 1
            else:
                updated += 1

        summary = {
            "created_entries": created,
            "updated_entries": updated,
            "skipped_rows": skipped,
            "skipped_wrong_month": skipped_wrong_month,
            "skipped_no_date": skipped_no_date,
            "sheet_name": sheet.title,
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "filename": upload.name,
            "target_month": f"{target_year:04d}-{target_month:02d}",
        }

        return render(
            request,
            self.template_name,
            {
                "sheet_names": sheet_names,
                "selected_sheet": sheet.title,
                "preview_headers": preview_headers,
                "preview_rows": preview_rows,
                "summary": summary,
            },
        )
