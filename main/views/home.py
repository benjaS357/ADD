import unicodedata
from decimal import Decimal, InvalidOperation

from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from ..models import Product, Pvp


def _normalize_header(name: str) -> str:
    decomposed = unicodedata.normalize("NFKD", name)
    ascii_only = decomposed.encode("ascii", "ignore").decode("ascii")
    return ascii_only.strip().lower()


def _index_map(headers):
    return {_normalize_header(h): idx for idx, h in enumerate(headers) if h}


def _value(row, idx):
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if isinstance(val, str):
        return val.strip()
    return val


@method_decorator(csrf_exempt, name="dispatch")
class HomeView(View):
    template_name = "upload_excel.html"

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        upload = request.FILES.get("file")
        if not upload:
            return render(request, self.template_name, {"error": "Selecciona un archivo .xlsx"})

        if not upload.name.lower().endswith(".xlsx"):
            return render(request, self.template_name, {"error": "El archivo debe ser .xlsx"})

        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception as exc:  # pragma: no cover - defensive
            return render(request, self.template_name, {"error": f"No se pudo leer el Excel: {exc}"})

        summary = {
            "products": {"created": 0, "updated": 0, "skipped": 0},
            "pvp": {"created": 0, "updated": 0, "skipped": 0, "missing_product": 0},
        }

        # Procesa Maestro de Productos
        if "Maestro de Productos" in workbook.sheetnames:
            sheet = workbook["Maestro de Productos"]
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                headers = rows[0]
                indexes = _index_map(headers)
                mapping = {
                    "code": indexes.get("cod articulo"),
                    "name": indexes.get("nom articulo"),
                    "group": indexes.get("grupo de articulos"),
                    "manufacturer": indexes.get("fabricante"),
                    "category": indexes.get("categoria"),
                    "subcategory": indexes.get("sub categoria"),
                    "size": indexes.get("tamano"),
                }
                for row in rows[1:]:
                    code = _value(row, mapping["code"])
                    name = _value(row, mapping["name"])
                    if not code or not name:
                        summary["products"]["skipped"] += 1
                        continue
                    defaults = {
                        "name": name,
                        "group": _value(row, mapping["group"]) or "",
                        "manufacturer": _value(row, mapping["manufacturer"]) or "",
                        "category": _value(row, mapping["category"]) or "",
                        "subcategory": _value(row, mapping["subcategory"]) or "",
                        "size": _value(row, mapping["size"]) or "",
                    }
                    obj, created = Product.objects.update_or_create(code=code, defaults=defaults)
                    if created:
                        summary["products"]["created"] += 1
                    else:
                        summary["products"]["updated"] += 1
        else:
            summary["products"]["skipped"] = "Hoja 'Maestro de Productos' no encontrada"

        # Procesa PVP
        if "PVP" in workbook.sheetnames:
            sheet = workbook["PVP"]
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                headers = rows[0]
                indexes = _index_map(headers)
                mapping = {
                    "sku": indexes.get("sku"),
                    "description": indexes.get("descripcion"),
                    "price": indexes.get("pvp"),
                }
                for row in rows[1:]:
                    sku = _value(row, mapping["sku"])
                    description = _value(row, mapping["description"]) or ""
                    price_raw = _value(row, mapping["price"])
                    if not sku:
                        summary["pvp"]["skipped"] += 1
                        continue
                    try:
                        price = Decimal(str(price_raw)) if price_raw not in (None, "") else Decimal("0")
                    except (InvalidOperation, TypeError):
                        summary["pvp"]["skipped"] += 1
                        continue
                    product = Product.objects.filter(code=sku).first()
                    if not product:
                        summary["pvp"]["missing_product"] += 1
                        continue
                    obj, created = Pvp.objects.update_or_create(
                        sku=sku,
                        defaults={"product": product, "description": description, "price": price},
                    )
                    if created:
                        summary["pvp"]["created"] += 1
                    else:
                        summary["pvp"]["updated"] += 1
        else:
            summary["pvp"]["skipped"] = "Hoja 'PVP' no encontrada"

        return render(request, self.template_name, {"summary": summary})
